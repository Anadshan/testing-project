import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstName": "rohit", "lastName": "shetty", "gender": "Male"},
                          {"firstName": "sanjoli", "lastName": "verma", "gender": "Female"}]

    @staticmethod  # to call the method with class name without object we use staticmethod and we don't need any self
    # parameter too ,self parameter required only when the method is non-static
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("//home//sanjoli//pythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
